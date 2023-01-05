from openpyxl import Workbook
from openpyxl import load_workbook
import openpyxl


wb = load_workbook(filename = 'test_sheet.xlsx')
sheet = wb['Mysheet1']

for cell in sheet['B']:
   print(cell.value)

print(sheet['A1'].value)

sheet.delete_rows(1,1)

for row in sheet.iter_rows():
    print(row)

wb.save('test_sheet.xlsx')


wb = Workbook()
ws = wb.active
ws.title = "hello"
wb.save('this.xlsx')

'''
maxr = sheet.max_row
maxc = sheet.max_column

print(maxr)
print(maxc)

print('A'+'1')
'''
