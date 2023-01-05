from openpyxl import Workbook
from openpyxl import load_workbook


def no_student_name():
    wb = load_workbook(filename = 'appointmentsReport.xlsx')
    ws = wb['appointmentsReport']

    maxr = ws.max_row
    empty = '-'

    for i in range(1,maxr+1):
        student = ws['D' + str(i)].value

        if student == empty:
            parent = ws['B' + str(i)].value
            ws['D' + str(i)].value = parent

    wb.save('appointmentsReport.xlsx')

no_student_name()
