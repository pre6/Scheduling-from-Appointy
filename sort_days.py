from openpyxl import Workbook
from openpyxl import load_workbook

'''
wb = load_workbook(filename = 'appointmentsReport.xlsx')
sheet = wb['appointmentsReport']
maxr = sheet.max_row
maxc = sheet.max_column
h = sheet.cell(1, 1).value

while h is not None:
'''

def make_new_with_date(name):
    wb = load_workbook(filename = 'appointmentsReport.xlsx')
    new = Workbook()

    ws = wb['appointmentsReport']
    new_sheet = new.active
    new_sheet.title = 'Day'

    maxr = ws.max_row
    maxc = ws.max_column

    date = ws['A2'].value #Selects the first date

    for i in range(1,maxr + 1):
        column = 'A'
        row = str(i)
        if ws[column + row].value == date:
            new_sheet['A' + str(i)].value = ws['A' + str(i)].value
            new_sheet['B' + str(i)].value = ws['B' + str(i)].value
            new_sheet['C' + str(i)].value = ws['C' + str(i)].value
            new_sheet['D' + str(i)].value = ws['D' + str(i)].value
            new_sheet['E' + str(i)].value = ws['E' + str(i)].value
            m = i

    #h = sheet.cell(1, 1).value
    new.save(name + '.xlsx')
    ws.delete_rows(2,m-1)
    wb.save('appointmentsReport.xlsx')

make_new_with_date('saturday')
make_new_with_date('thursday')
make_new_with_date('wednesday')
make_new_with_date('tuesday')
make_new_with_date('monday')
