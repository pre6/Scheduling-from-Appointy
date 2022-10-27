from openpyxl import Workbook
from openpyxl import load_workbook

def fix_intake(name):

    wb = load_workbook(filename = name+'.xlsx')
    ws = wb['Home']
    ws_1 = wb['Online']

    maxr = ws.max_row
    maxc = ws.max_column

    mr = ws_1.max_row
    mc = ws_1.max_column

    for r in range(1,maxr+1):

        y = ws['E'+str(r)].value
        y = y.replace('Intake form:\n','')
        y = y.replace('Student Name:','')
        y = y.replace('Student #2: ,\n','')
        y = y.replace('Student #2:','')
        y = y.replace('Student #3:','')
        y = y.replace(',\n','')

        ws['E'+str(r)].value = y

    for i in range(2,mr+1):

        x = ws_1['E'+str(i)].value
        x = x.replace('Intake form:\n','')
        x = x.replace('Student #2 Name (if applicable):','')
        x = x.replace('Student #3 Name (if applicable): ','')

        ws_1['E'+str(r)].value = x

    wb.save(name+'.xlsx')

fix_intake('tuesday')
