import pandas as pd
from datetime import datetime
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from copy import copy

# Load your student data
df = pd.read_csv("appointmentsReport.csv")  # columns: Date, Time, Student Name

# Drop unwanted columns
df.drop(columns=['Booking_Date', 'Booking_Time', 'Customer_Email', 'Customer_Address', 'Customer_Contact', 'Customer_Timezone', 'Booked_By', 'Service_Duration_mins', 'Amount', 'Staff_Name', 'Staff_Email', 'Resource', 'Customer_Remark', 'Appointment_Status', 'Arrival_Status', 'Admin_Note', 'Payment_Info', 'AddOn'], inplace=True)

# fix strings 
df['Intake_Form'] = df['Intake_Form'].str.replace(r'Intake form:\n|Student Name:|Student #2: ,\n|Student #2:|Student #3:|Student #2 Name \(if applicable\): ,|Student #3 Name \(if applicable\): |,\n|\n', '', regex=True).str.strip()
# print(df.head())

df['Intake_Form'] = df.apply(lambda row: row['Customer_Name'] if row['Intake_Form'] == '-' else row['Intake_Form'], axis=1)




# Convert Date and Time columns to datetime
# Convert date and time strings to datetime objects
df['Appointment_Date'] = pd.to_datetime(df['Appointment_Date'], format="%d %B %Y")
df['Appointment_Time'] = pd.to_datetime(df['Appointment_Time'], format="%I:%M %p").dt.time


# print(df['Appointment_Date'])


# print(df['Appointment_Time'])




# Add a weekday name column (e.g., 'Monday')
df['DayName'] = df['Appointment_Date'].dt.day_name()

# print(df['DayName'])





# Create 5-minute time slots from 10:00 AM to 8:00 PM
time_slots = pd.date_range("10:00", "20:00", freq="5min").time
template_df = pd.DataFrame({'Time': time_slots})

# Colors
hex_colors = ['FCE5CD', 'D9EAD3', '9FC5E8', 'EAD1DC', 'FFD966']

color_groups = 4  # every 4th column = same color/teacher

'''
Appointment_Date', 'Appointment_Time', 'Customer_Name', 'Service',
       'Intake_Form
'''



with pd.ExcelWriter("teacher_schedule_colored_reordered.xlsx", engine='openpyxl') as writer:
    for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
        day_df = df[df['DayName'] == day]
        # Create an initial schedule with the Time column and two columns for notes
        schedule = pd.DataFrame({'Time': time_slots})
        
        
        # schedule['Note 1'] = ''
        # schedule['Note 2'] = ''  # Add a second column for notes (besides the Time column)
        

        schedule = pd.DataFrame({'Time': time_slots})
        table_columns = []

        bookings = []  # list of (col_name, start_idx, length, group)

        for _, row in day_df.iterrows():
            student_time = row['Appointment_Time']
            student_name = row['Intake_Form']
            try:
                start_idx = schedule[schedule['Time'] == student_time].index[0]
            except IndexError:
                continue

            placed = False
            num_slots = 12

            for table_index in range(len(table_columns)):
                col = table_columns[table_index]
                if start_idx + num_slots <= len(schedule):
                    slot_range = schedule.loc[start_idx:start_idx+num_slots-1, col]
                    if (slot_range == '').all():
                        schedule.at[start_idx, col] = student_name
                        for i in range(1, num_slots):
                            schedule.at[start_idx + i, col] = 'BLOCKED'
                        group_id = table_index % color_groups
                        bookings.append((col, start_idx, num_slots, group_id))
                        placed = True
                        break

            if not placed:
                col = f'Col {len(table_columns) + 1}'
                table_columns.append(col)
                schedule[col] = ''
                schedule.at[start_idx, col] = student_name
                for i in range(1, num_slots):
                    if start_idx + i < len(schedule):
                        schedule.at[start_idx + i, col] = 'BLOCKED'
                group_id = (len(table_columns) - 1) % color_groups
                bookings.append((col, start_idx, num_slots, group_id))

        # Clean up blocked markers
        for col in table_columns:
            schedule[col] = schedule[col].replace('BLOCKED', '')

        # Reorder columns so same group (teacher) is together
        reordered_cols = ['Time']
        for group in range(color_groups):
            reordered_cols.extend([col for i, col in enumerate(table_columns) if i % color_groups == group])
        schedule = schedule[reordered_cols]

        schedule.to_excel(writer, sheet_name=day, index=False)
        
        
        
# Step 2: Apply color fills
wb = load_workbook("teacher_schedule_colored_reordered.xlsx")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    schedule_df = pd.read_excel("teacher_schedule_colored_reordered.xlsx", sheet_name=sheet_name)
    num_cols = schedule_df.shape[1]

    for col_idx in range(2, num_cols + 1):  # Skip Time column
        table_col_letter = ws.cell(row=1, column=col_idx).column_letter
        color_idx = ((col_idx - 2) // 4) % len(hex_colors)
        fill = PatternFill(start_color=hex_colors[color_idx], end_color=hex_colors[color_idx], fill_type='solid')

        row_idx = 2
        while row_idx <= len(time_slots) + 1:
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value and cell.value != '':
                # Color the current cell and next 11
                for i in range(12):
                    if row_idx + i <= len(time_slots) + 1:
                        ws.cell(row=row_idx + i, column=col_idx).fill = fill
                row_idx += 12
            else:
                row_idx += 1

# Save final workbook
wb.save("teacher_schedule_colored_reordered.xlsx")



      
        
#and uncolored

from openpyxl.utils import get_column_letter

sheets_to_remove = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    rows_to_delete = []
    
    for row_idx in range(2, max_row + 1):  # Skip header
        empty = True
        uncolored = True

        for col_idx in range(2, max_col + 1):  # Skip 'Time' column
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value not in (None, '', ' '):
                empty = False
            if cell.fill.start_color.rgb not in (None, '00000000', 'FFFFFFFF'):
                uncolored = False

        if empty and uncolored:
            rows_to_delete.append(row_idx)

    # Delete rows from bottom to top to avoid shifting
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx)

    # If sheet only has header left (or is totally empty), mark for deletion
    if ws.max_row <= 1:
        sheets_to_remove.append(sheet_name)

# Remove fully empty sheets
for sheet in sheets_to_remove:
    wb.remove(wb[sheet])
    
from openpyxl.utils import get_column_letter
from datetime import datetime, time

# Define minute sets
morning_minutes = {'20', '25', '35', '40', '45', '50', '55'}
evening_minutes = {'05', '10', '15', '20', '25', '50', '55'}

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    max_row = ws.max_row
    max_col = ws.max_column

    for row in range(max_row, 1, -1):  # skip header
        time_cell = ws.cell(row=row, column=1)
        time_value = time_cell.value
        
  

        # Skip if there's no valid time
        parsed_time = None
        if isinstance(time_value, (datetime, pd.Timestamp)):
            parsed_time = time_value.time()
        elif isinstance(time_value, str):
            try:
                parsed_time = datetime.strptime(time_value.strip(), "%H:%M:%S").time()
            except ValueError:
                continue

        if not parsed_time:
            continue

        hour, minute = parsed_time.hour, parsed_time.minute
        minute_str = f"{minute:02}"

        delete_row = False

        # Time-based rules
        if time(10, 0) <= parsed_time <= time(15, 0):  # 10:00 AM to 3:00 PM
            if minute_str in morning_minutes:
                delete_row = True
        elif time(15, 30) <= parsed_time <= time(20, 30):  # 3:30 PM to 8:30 PM
            if minute_str in evening_minutes:
                delete_row = True

        if delete_row:
            is_empty = True
            for col in range(2, max_col + 1):
                if ws.cell(row=row, column=col).value not in (None, '', 'BLOCKED'):
                    is_empty = False
                    break
            if is_empty:
                ws.delete_rows(row, 1)




# Save cleaned workbook
wb.save("teacher_schedule_colored_reordered.xlsx")

from openpyxl.styles import Font
from openpyxl.worksheet.copier import WorksheetCopy

# Create a new sheet for the consolidated data
if "All Days" in wb.sheetnames:
    del wb["All Days"]  # Delete if already exists
summary_ws = wb.create_sheet("All Days")

current_row = 1  # Track where to paste next

for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']:
    if day not in wb.sheetnames:
        continue
    day_ws = wb[day]
    max_row = day_ws.max_row
    max_col = day_ws.max_column

    # Optional: add a label for the day
    summary_ws.cell(row=current_row, column=1, value=day).font = Font(bold=True)
    current_row += 1

    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            source_cell = day_ws.cell(row=r, column=c)
            target_cell = summary_ws.cell(row=current_row, column=c)

            # Copy value
            target_cell.value = source_cell.value

            # Copy fill color
            if source_cell.fill and source_cell.fill.fill_type:
                target_cell.fill = copy(source_cell.fill)

            # Copy font and alignment (optional)
            target_cell.font = copy(source_cell.font)
            target_cell.alignment = copy(source_cell.alignment)

        current_row += 1

    current_row += 1  # Add a blank row between days
    
for sheet_name in wb.sheetnames:
    if sheet_name != "All Days":
        del wb[sheet_name]

# Save cleaned workbook
wb.save("teacher_schedule_colored_reordered.xlsx")