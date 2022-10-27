from openpyxl import Workbook
from openpyxl import load_workbook

def home_online(name):
    wb = load_workbook(filename = name +'.xlsx')
    ws_ic = wb.create_sheet("Home")
    ws_ol = wb.create_sheet("Online")
    ws = wb['Day']

    maxr = ws.max_row
    maxc = ws.max_column
    i = 1
    y = 1

    for r in range(1,maxr+1):
        if ws['D'+str(r)].value == 'In-Centre Session - 1 Hour(1) ':
            ws_ic['A' + str(i)].value = ws['A' + str(r)].value
            ws_ic['B' + str(i)].value = ws['B' + str(r)].value
            ws_ic['C' + str(i)].value = ws['C' + str(r)].value
            ws_ic['D' + str(i)].value = ws['D' + str(r)].value
            ws_ic['E' + str(i)].value = ws['E' + str(r)].value
            i += 1

        else:
            ws_ol['A' + str(y)].value = ws['A' + str(r)].value
            ws_ol['B' + str(y)].value = ws['B' + str(r)].value
            ws_ol['C' + str(y)].value = ws['C' + str(r)].value
            ws_ol['D' + str(y)].value = ws['D' + str(r)].value
            ws_ol['E' + str(y)].value = ws['E' + str(r)].value
            y += 1

    wb.save( name +'.xlsx')

home_online('tuesday')
