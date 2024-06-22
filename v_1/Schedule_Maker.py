from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from openpyxl.styles import PatternFill
import os
import tkinter as tk
from tkinter import filedialog


'''Part 1: Select file with Appointemnts and Format the file'''

# Global variable to store the selected file path
selected_file_path = ""

def browse_and_select_file():
    global selected_file_path
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if file_path:
        selected_file_path = file_path
        print(f"Selected file: {selected_file_path}")
    process_file()

def process_file():
    global selected_file_path
    if selected_file_path:
        file_path = selected_file_path
        delete_bad_columns(file_path)
        fix_intake(file_path)
        format_date_time(file_path)
        no_student_name(file_path)
        home_online(file_path)

def delete_bad_columns(file_path):
    ''' Deletes the columns that are uneeded '''

    wb = load_workbook(filename = file_path)
    sheet = wb.worksheets[0]
    #wb['appointmentsReport']
    sheet.delete_cols(3, 2)
    sheet.delete_cols(4, 4)
    sheet.delete_cols(5, 7)
    sheet.delete_cols(6, 5)
    sheet.delete_rows(1,1)
    wb.save(file_path)

def fix_intake(file_path):
    ''' The Students Names are in a weird formatted string it looks like:
    Inake Form:
        Student 1: Name
        Student 2: Name
    This fixes the format and removes everything except for the students names '''

    wb = load_workbook(filename = file_path)
    ws = wb.worksheets[0]
    #wb['appointmentsReport']

    maxr = ws.max_row
    maxc = ws.max_column

    for r in range(1,maxr+1):

        y = ws['E'+str(r)].value
        y = y.replace('Intake form:\n','')
        y = y.replace('Student Name:','')
        y = y.replace('Student #2: ,\n','')
        y = y.replace('Student #2:','')
        y = y.replace('Student #3:','')
        y = y.replace('Student #2 Name (if applicable): ,','')
        y = y.replace('Student #3 Name (if applicable): ','')
        y = y.replace(',\n','')
        y = y.replace('\n','')

        ws['E'+str(r)].value = y

    wb.save(file_path)

def format_date_time(file_path):
    '''Combines two columns together, One with the date and the other with the
    time. The date and time are in datetime format'''

    wb = load_workbook(filename = file_path)
    ws = wb.worksheets[0]
    #wb['appointmentsReport']

    ws.insert_cols(3)

    maxr = ws.max_row

    for i in range(1,maxr+1):
        column = 'A'
        row  = str(i)
        date = ws[column + row].value.date()
        time = ws['B'+row].value
        new = datetime.datetime.combine(date,time)

        ws['C' + row].value = new

    ws.delete_cols(1, 2)
    wb.save(file_path)
def no_student_name(file_path):
    '''Some parents do not put in their child's names. In that case this function
    takes in the parents name instead'''

    wb = load_workbook(filename = file_path)
    ws = wb.worksheets[0]
    #wb['appointmentsReport']

    maxr = ws.max_row
    empty = '-'

    for i in range(1,maxr+1):
        student = ws['D' + str(i)].value

        if student == empty:
            parent = ws['B' + str(i)].value
            ws['D' + str(i)].value = parent

    wb.save(file_path)

def home_online(file_path):
    ''' Some students are In person vs Online, This function storts it into two
    separate sheets '''

    wb = load_workbook(filename = file_path)
    ws_ic = wb.create_sheet("In_Center")
    ws_ol = wb.create_sheet("Online")
    ws = wb.worksheets[0]
    # wb['appointmentsReport']

    maxr = ws.max_row
    maxc = ws.max_column
    i = 1
    y = 1

    for r in range(1,maxr+1):
        if ws['C'+str(r)].value == 'In-Centre Session - 1 Hour(1) ':
            ws_ic['A' + str(i)].value = ws['A' + str(r)].value
            ws_ic['B' + str(i)].value = ws['B' + str(r)].value
            ws_ic['C' + str(i)].value = ws['C' + str(r)].value
            ws_ic['D' + str(i)].value = ws['D' + str(r)].value
            ws_ic['E' + str(i)].value = ws['E' + str(r)].value
            i += 1

        else:
            ws_ol['A' + str(y)].value = ws['A' + str(r)].value
            ws_ol['B' + str(y)].value = ws['B' + str(r)].value
            ws_ol['C' + str(y)].value = ws['C' + str(r)].value
            ws_ol['D' + str(y)].value = ws['D' + str(r)].value
            ws_ol['E' + str(y)].value = ws['E' + str(r)].value
            y += 1

    wb.save(file_path)


''' Part 2: Fix and select Template'''
selected_template_path = ""
def browse_and_select_template():
    global selected_template_path
    template_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if template_path:
        # Add code here to handle the selected template file
        print(f"Selected template: {template_path}")
        selected_template_path = template_path
    copy_template()
    create_schedule()

def copy_template():
    ''' Creates a copy of the template and creates a file called Appointments'''
    global selected_template_path

    wb = Workbook()
    wb.save(filename = 'Appointments.xlsx')
    wb_temp = load_workbook(filename = selected_template_path)
    wb = load_workbook('Appointments.xlsx')
    wb_temp.save("Appointments.xlsx")


'''Part 3: create the schedule'''
missing_students = []


def create_list(selected_file_path):
    ''' Creates a list of two lists. One line is for the online students the other
    is for the inperson students. Each list consists of dictionaries with students
    that will appear in each column for a particular day. In a dictionary, each
    student that follows is exactly one hour after each other:
    Example:
    [{datetime.datetime(2023, 1, 21, 10, 0): ['Student 1'],
    datetime.datetime(2023, 1, 21, 11, 0): [' Student 2 '],
    datetime.datetime(2023, 1, 21, 12, 0): [' Student 3 ']}]

    Two lists will be created, they will be two entries in a large list
    '''
    wb = load_workbook(filename = selected_file_path)
    ws_ic = wb.worksheets[1]
    #wb['In_Center']
    ws_ol = wb.worksheets[2]
    #wb['Online']

    lst_ic = []
    lst_ol = []
    maxr = ws_ic.max_row
    maxr_ol = ws_ol.max_row



    for r in range(1,maxr+1):
        dict = {}
        date = ws_ic['A'+str(r)].value
        dict[date] = [ws_ic['D' + str(r)].value]
        delta = datetime.timedelta(hours = 1)

        for i in range(r+1,maxr+1):
            next_date = ws_ic['A'+str(i)].value
            if next_date == None:
                break
            elif next_date == date + delta:
                dict[next_date] = [ws_ic['D' + str(i)].value]
                ws_ic.delete_rows(i,1)
                date = next_date
        if dict != {None:[None]}:
            lst_ic.append(dict)



    for t in range(1,maxr_ol+1):
        dict_ol = {}
        date = ws_ol['A'+str(t)].value
        dict_ol[date] = [ws_ol['D' + str(t)].value]
        delta = datetime.timedelta(hours = 1)

        for j in range(t+1,maxr_ol+1):
            next_date = ws_ol['A'+str(j)].value
            if next_date == None:
                break
            elif next_date == date + delta:
                dict_ol[next_date] = [ws_ol['D' + str(j)].value]
                ws_ol.delete_rows(j,1)
                date = next_date
        if dict_ol != {None:[None]}:
            lst_ol.append(dict_ol)

    return [lst_ic,lst_ol]

def split_list(lst):
    ''' Splits the list of dictionaries into days'''

    sat = []
    thurs = []
    wed = []
    tues = []
    mon = []

    day = []
    day.append(sat)
    day.append(thurs)
    day.append(wed)
    day.append(tues)
    day.append(mon)
    i = 0
    while len(lst) > 0:
        delete = lst.copy()
        date = list(lst[0].keys())[0].date()
        for item in lst:
            new = list(item.keys())[0].date()
            if new == date:
                day[i].append(item)
                delete.remove(item)
        lst = delete

        i += 1


    for empty in day:
        if empty == []:
            day.remove(empty)
    return day


def get_cells_ol(lst,day_num):
    ''' Gets the appropriate cell column and number for the first entry in
    each dictionary, these are for the ONLINE students. They take up the first 2
    columns '''

    Alpha = ['C','D','AA','AB']
    weekday = list(lst[0].keys())[0].weekday()

    if  (weekday == 5) or (weekday == 4):
        initial_time = datetime.time(hour = 10)
    else:
        initial_time = datetime.time(hour = 15, minute = 30)
    initial_column = 'C'

    if (weekday == 5) or (weekday == 4):
        initial_row = 118
    else:
        initial_row = 6+(weekday)*28

    for i in range(22):

        for column_dict in lst:

            start_time = list(column_dict.keys())[0]

            if start_time.time() == initial_time:
                student = column_dict[start_time]
                student.append(initial_column)
                student.append(initial_row)
                num = Alpha.index(initial_column) + 1
                initial_column = Alpha[num]

        if (weekday == 5) or (weekday == 4):
            initial_row = (118 + i + 1)
            if initial_time.minute == 15:
                delta = datetime.timedelta(minutes = 15)
            elif initial_time.minute == 30:
                delta = datetime.timedelta(minutes = 30)
            else:
                delta = datetime.timedelta(minutes = 5)
        else:
            initial_row = ((28 * (weekday)+6) + i + 1)
            if initial_time.minute == 45:
                delta = datetime.timedelta(minutes = 15)
            elif initial_time.minute == 0:
                delta = datetime.timedelta(minutes = 30)
            else:
                delta = datetime.timedelta(minutes = 5)

        rand_day = datetime.date(2021, 1, 4)
        new_date_with_time = datetime.datetime.combine(rand_day,initial_time)

        initial_time = (new_date_with_time + delta).time()
    return lst

def get_cell(lst,day_num):
    ''' Gets the appropriate cell column and number for the first entry in
    each dictionary,these are for the in center students'''


    Alpha = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY']
    weekday = list(lst[0].keys())[0].weekday()
    if  (weekday == 5) or (weekday == 4):
        initial_time = datetime.time(hour = 10)
    else:
        initial_time = datetime.time(hour = 15, minute = 30)
    initial_column = 'H'

    if (weekday == 5) or (weekday == 4):
        initial_row = 118
    else:
        initial_row = 6+(weekday)*28

    used_col = ['A','B','C','D','E','F','G']
    a = datetime.time(hour = 15, minute = 30)
    b = datetime.time(hour = 15, minute = 35)
    c = datetime.time(hour = 15, minute = 40)
    d = datetime.time(hour = 15, minute = 45)
    e = datetime.time(hour = 16)



    f = datetime.time(hour = 10)
    g = datetime.time(hour = 10, minute = 5)
    h = datetime.time(hour = 10, minute = 10)
    i = datetime.time(hour = 10, minute = 15)
    j = datetime.time(hour = 10, minute = 30)

    time_list = [a,b,c,d,e,f,g,h,i,j]

    for i in range(25):

        for column_dict in lst:

            start_time = list(column_dict.keys())[0]
            if start_time.time() == initial_time:

                if initial_time in time_list:
                    used_col.append(initial_column)

                    student = column_dict[start_time]
                    student.append(initial_column)
                    student.append(initial_row)

                    num = Alpha.index(initial_column) + 4
                    initial_column = Alpha[num]

                else:
                    used_col.append(initial_column)
                    student = column_dict[start_time]
                    student.append(initial_column)
                    student.append(initial_row)
                    num = Alpha.index(initial_column) + 1
                    initial_column = Alpha[num]

        rand_day = datetime.date(2000, 1, 4)

        if (weekday == 5) or (weekday==4):
            initial_row = (118 + i + 1)
            if initial_time.minute == 15:
                delta = datetime.timedelta(minutes = 15)
            elif initial_time.minute == 30:
                delta = datetime.timedelta(minutes = 30)
            else:
                delta = datetime.timedelta(minutes = 5)
        else:
            initial_row = ((28 * (weekday)+6) + i + 1)
            if initial_time.minute == 45:
                delta = datetime.timedelta(minutes = 15)
            elif initial_time.minute == 0:
                delta = datetime.timedelta(minutes = 30)
            else:
                delta = datetime.timedelta(minutes = 5)

        new_date_with_time = datetime.datetime.combine(rand_day,initial_time)
        initial_time = (new_date_with_time + delta).time()

        if initial_time in time_list:
            num = 7 + i + 1
            initial_column = Alpha[num]

        else:

            for n in used_col:
                if n in Alpha:
                    Alpha.remove(n)

            initial_column = Alpha[0]

    return lst


def check(lst):

    Alpha = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY']

    column_num = 'C'
    row_num = 146
    for item in lst:
        for dict in item:
            try:
                col = list(dict.values())[0][1]
                row = list(dict.values())[0][2]
            except:
                first_key = list(dict.keys())[0]
                student = dict[first_key]
                student.append(column_num)
                student.append(row_num)
                num = Alpha.index(column_num) + 1
                column_num = Alpha[num]
                missing_students.append(f"Missing Student: {student[0]}, Date: {first_key.date()}, Hour: {first_key.hour}, Minute: {first_key.minute}")


                # print('Missing Students:')
                # print('Student Name:', student[0])
                # print('Date:' , first_key.day)
                # print('Hour:', first_key.hour)
                # print('Minute', first_key.minute)
                # print()


    return lst




def fill_cell_ol(y):
    '''Fills the cells with the student name and colours, for online students'''

    wb = load_workbook(filename = 'Appointments.xlsx')
    ws = wb['Week']

    fill = PatternFill(start_color='EA9999', fill_type='solid')

    for day in y:
        for dict in day:
            col = list(dict.values())[0][1]
            row = list(dict.values())[0][2]

            for item in dict:
                student = dict[item][0]
                ws[col + str(row)].value =  student
                ws[col + str(row)].fill = fill
                ws[col + str(row+1)].fill = fill
                ws[col + str(row+2)].fill = fill
                ws[col + str(row+3)].fill = fill
                ws[col + str(row+4)].fill = fill
                row = row + 5

    wb.save('Appointments.xlsx')


def fill_cell(y):
    '''Fills the cells with the student name and colours, for inperson students'''

    wb = load_workbook(filename = 'Appointments.xlsx')
    ws = wb['Week']

    zeroth = PatternFill(start_color='FCE5CD', fill_type='solid')
    first = PatternFill(start_color='D9EAD3', fill_type='solid')
    second = PatternFill(start_color='9FC5E8', fill_type='solid')
    third = PatternFill(start_color='EAD1DC', fill_type='solid')
    fourth = PatternFill(start_color='FFD966', fill_type='solid')

    Alpha = ['H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI']
    colour_col = []
    number = 0
    interm = []
    for letter in Alpha:
        if number%4 ==0 and number != 0:
            colour_col.append(interm)
            interm = []
            interm.append(letter)
        else:
            interm.append(letter)
        number+=1

    d = 0

    for day in y:

        thing = list(day[0].keys())[0].date() #to set the date
        if (thing.weekday() == 5) or thing.weekday()==4:
            ws['A' + '117'].value = thing
        else:
            ws['A' + str(int(thing.weekday()*28+5))].value = thing


        for dict in day:
            col = list(dict.values())[0][1]
            row = list(dict.values())[0][2]

            if col in colour_col[0]:
                fill = zeroth
            elif col in colour_col[1]:
                fill = first
            elif col in colour_col[2]:
                fill = second
            elif col in colour_col[3]:
                fill = third
            else:
                fill = fourth

            for item in dict:
                student = dict[item][0]
                ws[col + str(row)].value =  student

                ws[col + str(row)].fill = fill
                ws[col + str(row+1)].fill = fill
                ws[col + str(row+2)].fill = fill
                ws[col + str(row+3)].fill = fill
                ws[col + str(row+4)].fill = fill
                row = row + 5
        d += 1
    wb.save('Appointments.xlsx')

def create_schedule():
    global selected_file_path
    global missing_students
    list_ol_ic = create_list(selected_file_path)
    list_ic_days = split_list(list_ol_ic[0])
    list_ol_days = split_list(list_ol_ic[1])

    finsihed_list_ol = []
    i = 0
    for item in list_ol_days:
        if item != []:
            f = get_cells_ol(item, i)
            finsihed_list_ol.append(f)
            i += 1

    finished_list_ic = []
    j = 0
    for thing in list_ic_days:
        if thing != []:
            f = get_cell(thing, j)
            finished_list_ic.append(f)
            j += 1

    check(finsihed_list_ol)
    check(finished_list_ic)

    fill_cell_ol(finsihed_list_ol)
    fill_cell(finished_list_ic)
    display_missing_students(missing_students)



def display_missing_students(missing_students):
    missing_students_frame = tk.Frame(app)
    missing_students_frame.grid(row=9, column=0, columnspan=2, padx=10, pady=10)

    # Create a label to indicate missing students
    label_missing_students = tk.Label(missing_students_frame, text="Missing Students:")
    label_missing_students.pack()

    # Create and populate a Text widget with missing student data
    text_missing_students = tk.Text(missing_students_frame, wrap=tk.WORD, height=10, width=60)  # Adjust the height and width here
    text_missing_students.pack()

    for student_info in missing_students:
        text_missing_students.insert(tk.END, student_info + '\n')





app = tk.Tk()
app.title("Schedule Creation App")

# Part 1: Prepare CSV Data
title_label_part1 = tk.Label(app, text="Part 1: Prepare CSV Data", font=("Helvetica", 14))
title_label_part1.grid(row=0, column=0, columnspan=2, padx=10, pady=10)

label1 = tk.Label(app, text="1. Have you downloaded the CSV file?")
label2 = tk.Label(app, text="2. Have you converted it to Excel?")
label3 = tk.Label(app, text="3. Browse and select the Excel file to clean:")

label1.grid(row=1, column=0, padx=10, pady=5, sticky="w")
label2.grid(row=2, column=0, padx=10, pady=5, sticky="w")

# Button to browse for the Excel file
label3.grid(row=3, column=0, padx=10, pady=5, sticky="w")
button_browse = tk.Button(app, text="Browse", command=browse_and_select_file)
button_browse.grid(row=3, column=1, padx=10, pady=5)
#
# # Button to process the selected file
# process_button = tk.Button(app, text="Process File", command=process_file)
# process_button.grid(row=4, column=1, padx=10, pady=10)

# Part 2: Create Template
title_label_part2 = tk.Label(app, text="Part 2: Create Template", font=("Helvetica", 14))
title_label_part2.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

# Label for making changes to the template
label_template = tk.Label(app, text="Make changes to template if needed")
label_template.grid(row=6, column=0, columnspan=2, pady=10)

# Button to browse for the template
button_template = tk.Button(app, text="Browse and Select Template", command=browse_and_select_template)
button_template.grid(row=7, column=0, columnspan=2, padx=10, pady=10)

# # Button to copy the template
# copy_template_button = tk.Button(app, text="Copy Template", command=copy_template)
# copy_template_button.grid(row=8, column=0, columnspan=2, padx=10, pady=10)
#
# #Button to create the final schedule
# create_schedule_button = tk.Button(app, text="Create Schedule", command=create_schedule)
# create_schedule_button.grid(row=9, column=0, columnspan=2, padx=10, pady=10)



app.mainloop()
