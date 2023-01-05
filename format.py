from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from openpyxl.styles import PatternFill
import os

def delete_files():
    file = 'appointmentsReport.csv'
    location = "C:/Users/Centre Director/Downloads/"
    path = os.path.join(location, file)
    os.remove(path)


def copy_template():
    wb = Workbook()
    wb.save('Appointments.xlsx')

    f ="C:/Users/Centre Director/Desktop/python_script_files/Template.xlsx"
    wb_temp = load_workbook(f)

    wb = load_workbook('Appointments.xlsx')

    wb_temp.save("Appointments.xlsx")

def delete_bad_columns():
    f = "C:/Users/Centre Director/Desktop/appointmentsReport.xlsx"
    wb = load_workbook(f)
    sheet = wb['appointmentsReport']
    sheet.delete_cols(3, 2)
    sheet.delete_cols(4, 4)
    sheet.delete_cols(5, 7)
    sheet.delete_cols(6, 5)
    sheet.delete_rows(1,1)
    wb.save(f)

def fix_intake():

    f = "C:/Users/Centre Director/Desktop/appointmentsReport.xlsx"
    wb = load_workbook(f)
    ws = wb['appointmentsReport']

    maxr = ws.max_row
    maxc = ws.max_column

    for r in range(1,maxr+1):

        y = ws['E'+str(r)].value
        y = y.replace('Intake form:\n','')
        y = y.replace('Student Name:','')
        y = y.replace('Student #2: ,\n','')
        y = y.replace('Student #2:','')
        y = y.replace('Student #3:','')
        y = y.replace('Student #2 Name (if applicable): ,','')
        y = y.replace('Student #3 Name (if applicable): ','')
        y = y.replace(',\n','')
        y = y.replace('\n','')

        ws['E'+str(r)].value = y

    wb.save(f)

def format_date_time():
    f = "C:/Users/Centre Director/Desktop/appointmentsReport.xlsx"
    wb = load_workbook(f)
    ws = wb['appointmentsReport']

    ws.insert_cols(3)

    maxr = ws.max_row

    for i in range(1,maxr+1):
        column = 'A'
        row  = str(i)
        date = ws[column + row].value.date()
        time = ws['B'+row].value
        new = datetime.datetime.combine(date,time)

        ws['C' + row].value = new

    ws.delete_cols(1, 2)
    wb.save(f)


def remove_students():
    f = "C:/Users/Centre Director/Desktop/python_script_files/highschool.xlsx"
    wb = load_workbook(f)
    ws = wb['Sheet1']

    maxr = ws.max_row

    highschool_students = []

    for i in range(1,maxr+1):
        highschool_students.append(ws['A'+str(i)].value)

    g = "C:/Users/Centre Director/Desktop/appointmentsReport.xlsx"

    wb_appoint = load_workbook(g)
    ws_appoint = wb_appoint['appointmentsReport']

    maxrow = ws_appoint.max_row

    delete = []

    for j in range(1,maxrow+1):
        student = ws_appoint['D' + str(j)].value
        for thing in highschool_students:
            if thing == student:
                delete.append(j)
    i = 0
    for thing in delete:

        ws_appoint.delete_rows(thing - i)
        i += 1


    wb_appoint.save(g)

def no_student_name():
    f = "C:/Users/Centre Director/Desktop/appointmentsReport.xlsx"
    wb = load_workbook(f)
    ws = wb['appointmentsReport']

    maxr = ws.max_row
    empty = '-'

    for i in range(1,maxr+1):
        student = ws['D' + str(i)].value

        if student == empty:
            parent = ws['B' + str(i)].value
            ws['D' + str(i)].value = parent

    wb.save(f)

def home_online():
    f = "C:/Users/Centre Director/Desktop/appointmentsReport.xlsx"
    wb = load_workbook(f)
    ws_ic = wb.create_sheet("In_Center")
    ws_ol = wb.create_sheet("Online")
    ws = wb['appointmentsReport']

    maxr = ws.max_row
    maxc = ws.max_column
    i = 1
    y = 1

    for r in range(1,maxr+1):
        if ws['C'+str(r)].value == 'In-Centre Session - 1 Hour(1) ':
            ws_ic['A' + str(i)].value = ws['A' + str(r)].value
            ws_ic['B' + str(i)].value = ws['B' + str(r)].value
            ws_ic['C' + str(i)].value = ws['C' + str(r)].value
            ws_ic['D' + str(i)].value = ws['D' + str(r)].value
            ws_ic['E' + str(i)].value = ws['E' + str(r)].value
            i += 1

        else:
            ws_ol['A' + str(y)].value = ws['A' + str(r)].value
            ws_ol['B' + str(y)].value = ws['B' + str(r)].value
            ws_ol['C' + str(y)].value = ws['C' + str(r)].value
            ws_ol['D' + str(y)].value = ws['D' + str(r)].value
            ws_ol['E' + str(y)].value = ws['E' + str(r)].value
            y += 1

    wb.save(f)

def create_list():
    f = "C:/Users/Centre Director/Desktop/appointmentsReport.xlsx"
    wb = load_workbook(f)
    ws_ic = wb['In_Center']
    ws_ol = wb['Online']

    lst_ic = []
    lst_ol = []
    maxr = ws_ic.max_row
    maxr_ol = ws_ol.max_row



    for r in range(1,maxr+1):
        dict = {}
        date = ws_ic['A'+str(r)].value
        dict[date] = [ws_ic['D' + str(r)].value]
        delta = datetime.timedelta(hours = 1)

        for i in range(r+1,maxr+1):
            next_date = ws_ic['A'+str(i)].value
            if next_date == None:
                break
            elif next_date == date + delta:
                dict[next_date] = [ws_ic['D' + str(i)].value]
                ws_ic.delete_rows(i,1)
                date = next_date
        if dict != {None:[None]}:
            lst_ic.append(dict)



    for t in range(1,maxr_ol+1):
        dict_ol = {}
        date = ws_ol['A'+str(t)].value
        dict_ol[date] = [ws_ol['D' + str(t)].value]
        delta = datetime.timedelta(hours = 1)

        for j in range(t+1,maxr_ol+1):
            next_date = ws_ol['A'+str(j)].value
            if next_date == None:
                break
            elif next_date == date + delta:
                dict_ol[next_date] = [ws_ol['D' + str(j)].value]
                ws_ol.delete_rows(j,1)
                date = next_date
        if dict_ol != {None:[None]}:
            lst_ol.append(dict_ol)

    return [lst_ic,lst_ol]

def split_list(lst):
    sat = []
    thurs = []
    wed = []
    tues = []
    mon = []

    day = []
    day.append(sat)
    day.append(thurs)
    day.append(wed)
    day.append(tues)
    day.append(mon)
    i = 0
    while len(lst) > 0:
        delete = lst.copy()
        date = list(lst[0].keys())[0].date()
        for item in lst:
            new = list(item.keys())[0].date()
            if new == date:
                day[i].append(item)
                delete.remove(item)
        lst = delete

        i += 1


    for empty in day:
        if empty == []:
            day.remove(empty)
    return day


def get_cells_ol(lst,day_num):
    Alpha = ['C','D','E','F','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG']
    weekday = list(lst[0].keys())[0].weekday()
    print(weekday)
    if  weekday == 5:
        initial_time = datetime.time(hour = 10)
    else:
        initial_time = datetime.time(hour = 15, minute = 30)
    initial_column = 'C'
    if weekday == 5:
        initial_row = 6+(weekday-1)*28
    else:
        initial_row = 6+(weekday)*28

    for i in range(22):

        for column_dict in lst:

            start_time = list(column_dict.keys())[0]

            if start_time.time() == initial_time:
                student = column_dict[start_time]
                student.append(initial_column)
                student.append(initial_row)
                num = Alpha.index(initial_column) + 1
                initial_column = Alpha[num]

        if weekday == 5:
            initial_row = ((28 * (weekday-1)+6) + i + 1)
        else:
            initial_row = ((28 * (weekday)+6) + i + 1)
        rand_day = datetime.date(2021, 1, 4)
        delta = datetime.timedelta(minutes = 15)
        new_date_with_time = datetime.datetime.combine(rand_day,initial_time)

        initial_time = (new_date_with_time + delta).time()
    return lst


def get_cell(lst,day_num):

    Alpha = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI']
    weekday = list(lst[0].keys())[0].weekday()
    if  weekday == 5:
        initial_time = datetime.time(hour = 10)
    else:
        initial_time = datetime.time(hour = 15, minute = 30)
    initial_column = 'G'

    if weekday == 5:
        initial_row = 6+(weekday-1)*28
    else:
        initial_row = 6+(weekday)*28

    used_col = ['A','B','C','D','E','F']
    a = datetime.time(hour = 15, minute = 30)
    b = datetime.time(hour = 15, minute = 45)
    c = datetime.time(hour = 16)
    d = datetime.time(hour = 10)
    e = datetime.time(hour = 10, minute = 15)
    f = datetime.time(hour = 10, minute = 30)

    time_list = [a,b,c,d,e,f]

    for i in range(17):

        for column_dict in lst:

            start_time = list(column_dict.keys())[0]
            if start_time.time() == initial_time:

                if initial_time in time_list:
                    used_col.append(initial_column)

                    student = column_dict[start_time]
                    student.append(initial_column)
                    student.append(initial_row)

                    num = Alpha.index(initial_column) + 3
                    initial_column = Alpha[num]

                else:
                    used_col.append(initial_column)
                    student = column_dict[start_time]
                    student.append(initial_column)
                    student.append(initial_row)
                    num = Alpha.index(initial_column) + 1
                    initial_column = Alpha[num]

        if weekday == 5:
            initial_row = ((28 * (weekday-1)+6) + i + 1)
        else:
            initial_row = ((28 * (weekday)+6) + i + 1)
        rand_day = datetime.date(2021, 1, 4)
        delta = datetime.timedelta(minutes = 15)
        new_date_with_time = datetime.datetime.combine(rand_day,initial_time)

        initial_time = (new_date_with_time + delta).time()

        if initial_time in time_list:
            num = Alpha.index('G') + i + 1
            initial_column = Alpha[num]

        else:

            for n in used_col:
                if n in Alpha:
                    Alpha.remove(n)

            initial_column = Alpha[0]

    return lst


def fill_cell_ol(y):

    wb = load_workbook(filename = 'Appointments.xlsx')
    ws = wb['Week']

    fill = PatternFill(start_color='EA9999', fill_type='solid')

    for day in y:
        for dict in day:
            col = list(dict.values())[0][1]
            row = list(dict.values())[0][2]

            for item in dict:
                student = dict[item][0]
                ws[col + str(row)].value =  student
                ws[col + str(row)].fill = fill
                ws[col + str(row+1)].fill = fill
                ws[col + str(row+2)].fill = fill
                ws[col + str(row+3)].fill = fill
                row = row + 4

    wb.save('Appointments.xlsx')


def fill_cell(y):
    wb = load_workbook(filename = 'Appointments.xlsx')
    ws = wb['Week']

    zeroth = PatternFill(start_color='FCE5CD', fill_type='solid')
    first = PatternFill(start_color='D9EAD3', fill_type='solid')
    second = PatternFill(start_color='9FC5E8', fill_type='solid')
    third = PatternFill(start_color='EAD1DC', fill_type='solid')
    fourth = PatternFill(start_color='FFD966', fill_type='solid')

    Alpha = ['G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI']
    colour_col = []
    number = 0
    interm = []
    for letter in Alpha:
        if number%3 ==0 and number != 0:
            colour_col.append(interm)
            interm = []
            interm.append(letter)
        else:
            interm.append(letter)
        number+=1

    d = 0

    for day in y:

        thing = list(day[0].keys())[0].date() #to set the date
        the_col = 117-28*d
        ws['A' + str(the_col)].value = thing

        for dict in day:
            col = list(dict.values())[0][1]
            row = list(dict.values())[0][2]

            if col in colour_col[0]:
                fill = zeroth
            elif col in colour_col[1]:
                fill = first
            elif col in colour_col[2]:
                fill = second
            elif col in colour_col[3]:
                fill = third
            else:
                fill = fourth

            for item in dict:
                student = dict[item][0]
                ws[col + str(row)].value =  student

                ws[col + str(row)].fill = fill
                ws[col + str(row+1)].fill = fill
                ws[col + str(row+2)].fill = fill
                ws[col + str(row+3)].fill = fill
                row = row + 4
        d += 1
    wb.save('Appointments.xlsx')

def delete_files_last():
    file = 'appointmentsReport.xlsx'
    location = "C:/Users/Centre Director/Desktop/"
    path = os.path.join(location, file)
    os.remove(path)



delete_files()
copy_template()
delete_bad_columns()
fix_intake()
format_date_time()
remove_students()
no_student_name()
home_online()


list_ol_ic = create_list()


list_ic_days = split_list(list_ol_ic[0])

list_ol_days = split_list(list_ol_ic[1])

finsihed_list_ol = []
i = 0
for item in list_ol_days:
    if item != []:
        f = get_cells_ol(item,i)
        finsihed_list_ol.append(f)
        i += 1


finished_list_ic = []
j = 0
for thing in list_ic_days:
    if thing != []:
        f = get_cell(thing,j)
        finished_list_ic.append(f)
        j += 1

fill_cell_ol(finsihed_list_ol)
fill_cell(finished_list_ic)

delete_files_last()
