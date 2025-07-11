from __future__ import print_function

from openpyxl import Workbook
from openpyxl import load_workbook
import datetime
from openpyxl.styles import PatternFill
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import Text, filedialog
import pandas as pd
# from datetime import datetime, time
import datetime

import pygsheets
import gspread
from googleapiclient.discovery import build
from google.oauth2 import service_account
from googleapiclient.http import MediaFileUpload
import google.auth
from googleapiclient.errors import HttpError
import json





'''Global variable is the date of next Monday'''



def read_config_file(file_path):
    config_dict = {}
    try:
        with open(file_path, 'r') as file:
            for line in file:
                line = line.strip()
                if line:
                    key, value = line.split(": ", 1)
                    config_dict[key] = value
        return config_dict
    except FileNotFoundError:
        return "File not found."



'''Part 1. Cleaning and formating the source csv file'''

def convert(file_path,save_file):
    '''1. convert csv file to excel using pandas'''
    df = pd.read_csv(file_path)
    excel_file = save_file
    df.to_excel(excel_file, index=False)

def date_time_convert(file_path):
    '''2. Merge the date time fields to one columns as datetimetype'''

    wb = load_workbook(file_path)
    ws = wb.active
    datetime_format = "%d %b %Y %I:%M %p"

    # Iterate through rows and convert Date and Time columns to datetime
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=2):
        date_value = row[0].value
        time_value = row[1].value

        if date_value is not None and time_value is not None:
            # Otherwise, parse the date and time from string format
            date_string = str(date_value)
            time_string = str(time_value)
            datetime_string = date_string + ' ' + time_string
            datetime_value = datetime.datetime.strptime(datetime_string, datetime_format)


            # Assign the datetime value back to the Date column
            row[0].value = datetime_value

    # delete the time column as it is not needed
    ws.delete_cols(2)

    # Save the updated Excel file
    wb.save(file_path)

    # Close the workbook
    wb.close()

def delete_bad_columns(file_path):
    ''' Deletes the columns that are uneeded '''

    wb = load_workbook(filename = file_path)
    sheet = wb.worksheets[0]
    
    sheet.delete_cols(18, 25)
    sheet.delete_cols(10, 7)
    sheet.delete_cols(5, 4)
    sheet.delete_cols(2, 2)
    sheet.delete_rows(1,1)
    wb.save(file_path)

def fix_intake(file_path):
    ''' The Students Names are in a weird formatted string it looks like:
    Inake Form:
        Student 1: Name
        Student 2: Name
    This fixes the format and removes everything except for the students names '''

    wb = load_workbook(filename = file_path)
    ws = wb.worksheets[0]

    maxr = ws.max_row
    maxc = ws.max_column

    for r in range(1,maxr+1):

        y = ws['D'+str(r)].value
        y = y.replace('Intake form:\n','')
        y = y.replace('Student Name:','')
        y = y.replace('Student #2: ,\n','')
        y = y.replace('Student #2:','')
        y = y.replace('Student #3:','')
        y = y.replace('Student #2 Name (if applicable): ,','')
        y = y.replace('Student #3 Name (if applicable): ','')
        y = y.replace(',\n','')
        y = y.replace('\n','')

        ws['D'+str(r)].value = y

    wb.save(file_path)

def no_student_name(file_path):
    '''Some parents do not put in their child's names. In that case this function
    takes in the parents name instead'''

    wb = load_workbook(filename = file_path)
    ws = wb.worksheets[0]

    maxr = ws.max_row
    empty = '-'

    for i in range(1,maxr+1):
        student = ws['D' + str(i)].value

        if student == empty:
            parent = ws['B' + str(i)].value
            ws['D' + str(i)].value = parent

    wb.save(file_path)

def home_online(file_path):
    ''' Some students are In person vs Online, This function storts it into two
    separate sheets '''

    wb = load_workbook(filename = file_path)
    ws_ic = wb.create_sheet("In_Center")
    ws_ol = wb.create_sheet("Online")
    ws = wb.worksheets[0]

    maxr = ws.max_row
    maxc = ws.max_column
    i = 1
    y = 1

    for r in range(1,maxr+1):
        if ws['C'+str(r)].value == 'In-Centre Session - 1 Hour(1) ':
            ws_ic['A' + str(i)].value = ws['A' + str(r)].value
            ws_ic['B' + str(i)].value = ws['B' + str(r)].value
            ws_ic['C' + str(i)].value = ws['C' + str(r)].value
            ws_ic['D' + str(i)].value = ws['D' + str(r)].value
            ws_ic['E' + str(i)].value = ws['E' + str(r)].value
            i += 1

        else:
            ws_ol['A' + str(y)].value = ws['A' + str(r)].value
            ws_ol['B' + str(y)].value = ws['B' + str(r)].value
            ws_ol['C' + str(y)].value = ws['C' + str(r)].value
            ws_ol['D' + str(y)].value = ws['D' + str(r)].value
            ws_ol['E' + str(y)].value = ws['E' + str(r)].value
            y += 1

    wb.save(file_path)


''' Part 2: Create Dictionaries and excel sheet'''

missing_students = []


def copy_template(save_path,selected_template_path):
    ''' Creates a copy of the template and creates a file called Appointments'''

    wb = Workbook()
    wb.save(save_path)
    wb_temp = load_workbook(filename = selected_template_path)
    wb = load_workbook(save_path)
    wb_temp.save(save_path)


def create_list(selected_file_path):
    ''' Creates a list of two lists. One line is for the online students the other
    is for the inperson students. Each list consists of dictionaries with students
    that will appear in each column for a particular day. In a dictionary, each
    student that follows is exactly one hour after each other:
    Example:
    [{datetime.datetime(2023, 1, 21, 10, 0): ['Student 1'],
    datetime.datetime(2023, 1, 21, 11, 0): [' Student 2 '],
    datetime.datetime(2023, 1, 21, 12, 0): [' Student 3 ']}]

    Two lists will be created, they will be two entries in a large list
    '''
    wb = load_workbook(filename = selected_file_path)
    ws_ic = wb.worksheets[1]
    #wb['In_Center']
    ws_ol = wb.worksheets[2]
    #wb['Online']

    lst_ic = []
    lst_ol = []
    maxr = ws_ic.max_row
    maxr_ol = ws_ol.max_row



    for r in range(1,maxr+1):
        dict = {}
        date = ws_ic['A'+str(r)].value
        dict[date] = [ws_ic['D' + str(r)].value]
        delta = datetime.timedelta(hours = 1)

        for i in range(r+1,maxr+1):
            next_date = ws_ic['A'+str(i)].value
            if next_date == None:
                break
            elif next_date == date + delta:
                dict[next_date] = [ws_ic['D' + str(i)].value]
                ws_ic.delete_rows(i,1)
                date = next_date
        if dict != {None:[None]}:
            lst_ic.append(dict)



    for t in range(1,maxr_ol+1):
        dict_ol = {}
        date = ws_ol['A'+str(t)].value
        dict_ol[date] = [ws_ol['D' + str(t)].value]
        delta = datetime.timedelta(hours = 1)

        for j in range(t+1,maxr_ol+1):
            next_date = ws_ol['A'+str(j)].value
            if next_date == None:
                break
            elif next_date == date + delta:
                dict_ol[next_date] = [ws_ol['D' + str(j)].value]
                ws_ol.delete_rows(j,1)
                date = next_date
        if dict_ol != {None:[None]}:
            lst_ol.append(dict_ol)

    return [lst_ic,lst_ol]

def split_list(lst):
    ''' Splits the list of dictionaries into days'''

    sat = []
    thurs = []
    wed = []
    tues = []
    mon = []

    day = []
    day.append(sat)
    day.append(thurs)
    day.append(wed)
    day.append(tues)
    day.append(mon)
    i = 0
    while len(lst) > 0:
        delete = lst.copy()
        date = list(lst[0].keys())[0].date()
        for item in lst:
            new = list(item.keys())[0].date()
            if new == date:
                day[i].append(item)
                delete.remove(item)
        lst = delete

        i += 1


    for empty in day:
        if empty == []:
            day.remove(empty)
    return day

def get_cells_ol(lst,day_num):
    ''' Gets the appropriate cell column and number for the first entry in
    each dictionary, these are for the ONLINE students. They take up the first 2
    columns '''

    Alpha = ['C','D','AA','AB','BA','BB','BC']
    weekday = list(lst[0].keys())[0].weekday()

    if  (weekday == 5) or (weekday == 4):
        initial_time = datetime.time(hour = 10)
    else:
        initial_time = datetime.time(hour = 15, minute = 30)
    initial_column = 'C'

    if (weekday == 5) or (weekday == 4):
        initial_row = 118
    else:
        initial_row = 6+(weekday)*28

    for i in range(22):

        for column_dict in lst:

            start_time = list(column_dict.keys())[0]

            if start_time.time() == initial_time:
                student = column_dict[start_time]
                student.append(initial_column)
                student.append(initial_row)
                num = Alpha.index(initial_column) + 1
                initial_column = Alpha[num]

        if (weekday == 5) or (weekday == 4):
            initial_row = (118 + i + 1)
            if initial_time.minute == 15:
                delta = datetime.timedelta(minutes = 15)
            elif initial_time.minute == 30:
                delta = datetime.timedelta(minutes = 30)
            else:
                delta = datetime.timedelta(minutes = 5)
        else:
            initial_row = ((28 * (weekday)+6) + i + 1)
            if initial_time.minute == 45:
                delta = datetime.timedelta(minutes = 15)
            elif initial_time.minute == 0:
                delta = datetime.timedelta(minutes = 30)
            else:
                delta = datetime.timedelta(minutes = 5)

        rand_day = datetime.date(2021, 1, 4)
        new_date_with_time = datetime.datetime.combine(rand_day,initial_time)

        initial_time = (new_date_with_time + delta).time()
    return lst

def get_cell(lst,day_num):
    ''' Gets the appropriate cell column and number for the first entry in
    each dictionary,these are for the in center students'''


    Alpha = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U',
             'V','W','X','Y','Z','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO'
             ,'AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY','AZ','BA','BB','BC','BD','BE','BF','BG']
    weekday = list(lst[0].keys())[0].weekday()
    if  (weekday == 5) or (weekday == 4):
        initial_time = datetime.time(hour = 10)
    else:
        initial_time = datetime.time(hour = 15, minute = 30)
    initial_column = 'H'

    if (weekday == 5) or (weekday == 4):
        initial_row = 118
    else:
        initial_row = 6+(weekday)*28

    used_col = ['A','B','C','D','E','F','G']
    a = datetime.time(hour = 15, minute = 30)
    b = datetime.time(hour = 15, minute = 35)
    c = datetime.time(hour = 15, minute = 40)
    d = datetime.time(hour = 15, minute = 45)
    e = datetime.time(hour = 16)



    f = datetime.time(hour = 10)
    g = datetime.time(hour = 10, minute = 5)
    h = datetime.time(hour = 10, minute = 10)
    i = datetime.time(hour = 10, minute = 15)
    j = datetime.time(hour = 10, minute = 30)

    time_list = [a,b,c,d,e,f,g,h,i,j]

    for i in range(25):

        for column_dict in lst:

            start_time = list(column_dict.keys())[0]
            if start_time.time() == initial_time:

                if initial_time in time_list:
                    used_col.append(initial_column)

                    student = column_dict[start_time]
                    student.append(initial_column)
                    student.append(initial_row)

                    num = Alpha.index(initial_column) + 4
                    initial_column = Alpha[num]

                else:
                    used_col.append(initial_column)
                    student = column_dict[start_time]
                    student.append(initial_column)
                    student.append(initial_row)
                    num = Alpha.index(initial_column) + 1
                    initial_column = Alpha[num]

        rand_day = datetime.date(2000, 1, 4)

        if (weekday == 5) or (weekday==4):
            initial_row = (118 + i + 1)
            if initial_time.minute == 15:
                delta = datetime.timedelta(minutes = 15)
            elif initial_time.minute == 30:
                delta = datetime.timedelta(minutes = 30)
            else:
                delta = datetime.timedelta(minutes = 5)
        else:
            initial_row = ((28 * (weekday)+6) + i + 1)
            if initial_time.minute == 45:
                delta = datetime.timedelta(minutes = 15)
            elif initial_time.minute == 0:
                delta = datetime.timedelta(minutes = 30)
            else:
                delta = datetime.timedelta(minutes = 5)

        new_date_with_time = datetime.datetime.combine(rand_day,initial_time)
        initial_time = (new_date_with_time + delta).time()

        if initial_time in time_list:
            num = 7 + i + 1
            initial_column = Alpha[num]

        else:

            for n in used_col:
                if n in Alpha:
                    Alpha.remove(n)

            initial_column = Alpha[0]

    return lst

def check(lst):

    Alpha = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI','AJ','AK','AL','AM','AN','AO','AP','AQ','AR','AS','AT','AU','AV','AW','AX','AY']

    column_num = 'C'
    row_num = 146
    for item in lst:
        for dict in item:
            try:
                col = list(dict.values())[0][1]
                row = list(dict.values())[0][2]
            except:
                first_key = list(dict.keys())[0]
                student = dict[first_key]
                student.append(column_num)
                student.append(row_num)
                num = Alpha.index(column_num) + 1
                column_num = Alpha[num]
                missing_students.append(f"Missing Student: {student[0]}, Date: {first_key.date()}, Hour: {first_key.hour}, Minute: {first_key.minute}")


                # print('Missing Students:')
                # print('Student Name:', student[0])
                # print('Date:' , first_key.day)
                # print('Hour:', first_key.hour)
                # print('Minute', first_key.minute)
                # print()


    return lst

def fill_cell_ol(y,file_path):
    '''Fills the cells with the student name and colours, for online students'''

    wb = load_workbook(file_path)
    ws = wb['Week']

    fill = PatternFill(start_color='EA9999', fill_type='solid')

    for day in y:
        for dict in day:
            col = list(dict.values())[0][1]
            row = list(dict.values())[0][2]

            for item in dict:
                student = dict[item][0]
                ws[col + str(row)].value =  student
                ws[col + str(row)].fill = fill
                ws[col + str(row+1)].fill = fill
                ws[col + str(row+2)].fill = fill
                ws[col + str(row+3)].fill = fill
                ws[col + str(row+4)].fill = fill
                row = row + 5

    wb.save(file_path)
    wb.close()

def fill_cell(y,file_path):
    '''Fills the cells with the student name and colours, for inperson students'''

    wb = load_workbook(file_path)
    ws = wb['Week']

    zeroth = PatternFill(start_color='FCE5CD', fill_type='solid')
    first = PatternFill(start_color='D9EAD3', fill_type='solid')
    second = PatternFill(start_color='9FC5E8', fill_type='solid')
    third = PatternFill(start_color='EAD1DC', fill_type='solid')
    fourth = PatternFill(start_color='FFD966', fill_type='solid')

    Alpha = ['H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V','W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG','AH','AI']
    colour_col = []
    number = 0
    interm = []
    for letter in Alpha:
        if number%4 ==0 and number != 0:
            colour_col.append(interm)
            interm = []
            interm.append(letter)
        else:
            interm.append(letter)
        number+=1

    d = 0

    for day in y:

        thing = list(day[0].keys())[0].date() #to set the date
        if (thing.weekday() == 5) or thing.weekday()==4:
            ws['A' + '117'].value = thing
        else:
            ws['A' + str(int(thing.weekday()*28+5))].value = thing


        for dict in day:
            col = list(dict.values())[0][1]
            row = list(dict.values())[0][2]

            if col in colour_col[0]:
                fill = zeroth
            elif col in colour_col[1]:
                fill = first
            elif col in colour_col[2]:
                fill = second
            elif col in colour_col[3]:
                fill = third
            else:
                fill = fourth

            for item in dict:
                student = dict[item][0]
                ws[col + str(row)].value =  student

                ws[col + str(row)].fill = fill
                ws[col + str(row+1)].fill = fill
                ws[col + str(row+2)].fill = fill
                ws[col + str(row+3)].fill = fill
                ws[col + str(row+4)].fill = fill
                row = row + 5
        d += 1
    wb.save(file_path)
    wb.close()

def create_schedule(selected_file_path,file_path):
    global missing_students
    list_ol_ic = create_list(selected_file_path)
    list_ic_days = split_list(list_ol_ic[0])
    list_ol_days = split_list(list_ol_ic[1])

    finsihed_list_ol = []
    i = 0
    for item in list_ol_days:
        if item != []:
            f = get_cells_ol(item, i)
            finsihed_list_ol.append(f)
            i += 1

    finished_list_ic = []
    j = 0
    for thing in list_ic_days:
        if thing != []:
            f = get_cell(thing, j)
            finished_list_ic.append(f)
            j += 1

    check(finsihed_list_ol)
    check(finished_list_ic)

    fill_cell_ol(finsihed_list_ol,file_path)
    fill_cell(finished_list_ic,file_path)





def delete_file_excel(file_path):
    # Check if the file exists before attempting to delete it
    if os.path.exists(file_path):
        # Delete the file
        os.remove(file_path)
        print(f"{file_path} has been deleted.")
    else:
        print(f"{file_path} does not exist, so it cannot be deleted.")


'''Part 3 copy to google drive'''


def copy_excel_to_sheets(service_account_key_file,folder_id,local_excel_path,monday_date):
    # Define the desired scopes
    scopes = ['https://www.googleapis.com/auth/drive']

    # Create credentials
    credentials = service_account.Credentials.from_service_account_file(
        service_account_key_file, scopes=scopes
    )

    # Authenticate with Google Drive API
    drive_service = build('drive', 'v3', credentials=credentials)


    # Upload the Excel file to Google Drive
    file_metadata = {
        'name': monday_date+'.xlsx',
        'parents': [folder_id]
    }

    media = MediaFileUpload(local_excel_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    file = drive_service.files().create(
        body=file_metadata,
        media_body=media,
    ).execute()
    file_id = file.get('id')
    print(f'New file ID: {file_id}')

    # Return the file ID if needed
    return file_id

def convert_excel_to_sheets(service_account_key_file,new_file_id,monday_date):
    # Define the desired scopes
    scopes = ['https://www.googleapis.com/auth/drive']

    # Create credentials
    credentials = service_account.Credentials.from_service_account_file(
        service_account_key_file, scopes=scopes
    )

    # Initialize the Drive API service
    drive_service = build('drive', 'v3',credentials=credentials)

    # Define the MIME type for Google Sheets
    sheets_mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'  # MIME type for .xlsx files

    # Create a copy of the file with the Google Sheets MIME type
    new_metadata = {
        'mimeType': 'application/vnd.google-apps.spreadsheet',  # Google Sheets MIME type
        'name': monday_date  # Name for the converted Google Sheet
    }

    converted_file = drive_service.files().copy(fileId=new_file_id, body=new_metadata).execute()

    # Now, you can update or modify the converted file as needed

    # Example: Rename the converted file
    drive_service.files().update(fileId=converted_file['id'], body={'name': monday_date}).execute()

    file_id = converted_file.get('id')
    print(f'New file ID: {file_id}')

    print('File converted and updated:', converted_file['name'])

    return file_id

def delete_file(service_account_key_file, file_id):
    # Define the desired scopes
    scopes = ['https://www.googleapis.com/auth/drive']

    # Create credentials
    credentials = service_account.Credentials.from_service_account_file(
        service_account_key_file, scopes=scopes
    )

    # Authenticate with Google Drive API
    drive_service = build('drive', 'v3', credentials=credentials)

    try:
        # Delete the file with the specified file ID
        drive_service.files().delete(fileId=file_id).execute()
        print(f"File with ID {file_id} has been deleted.")
    except Exception as e:
        print(f"An error occurred: {e}")

def copy_column_a(service_account_key_file,source_spreadsheet_id,destination_spreadsheet_id):
    # Authenticate with the Google Sheets API using credentials
    scopes = ['https://www.googleapis.com/auth/spreadsheets']
    creds = service_account.Credentials.from_service_account_file(
        service_account_key_file,
        scopes=scopes
    )

    gc = gspread.authorize(creds)

    # Access the specific worksheets within the spreadsheets
    source_worksheet = gc.open_by_key(source_spreadsheet_id).worksheet('Week')  # Adjust the sheet name
    destination_worksheet = gc.open_by_key(destination_spreadsheet_id).worksheet('Week')  # Adjust the sheet name

    # Get the values from the source column A
    source_column = source_worksheet.col_values(1)

    # print([[value] for value in source_column])

    # Write the values to the destination column A starting from A1
    destination_worksheet.update([[value] for value in source_column],'A1:A1000')


def write_values_to_sheet(service_account_key_file, spreadsheet_id,next_monday):
    # Authenticate with Google Sheets API using the service account credentials
    credentials = service_account.Credentials.from_service_account_file(
        service_account_key_file, scopes=['https://www.googleapis.com/auth/spreadsheets']
    )

    service = build('sheets', 'v4', credentials=credentials)
    cell_locations = ['A5', 'A33', 'A61', 'A89', 'A117']
    values = [next_monday.strftime('%Y-%m-%d'), (next_monday + datetime.timedelta(days=1)).strftime('%Y-%m-%d'),
                (next_monday + datetime.timedelta(days=2)).strftime('%Y-%m-%d'),
                (next_monday + datetime.timedelta(days=3)).strftime('%Y-%m-%d'),
                (next_monday + datetime.timedelta(days=5)).strftime('%Y-%m-%d')]

    try:
        # Define the data to update
        value_input_option = 'RAW'
        data = [{'range': location, 'values': [[value]]} for location, value in zip(cell_locations, values)]
        body = {'valueInputOption': value_input_option, 'data': data}

        # Update the values in the specified cells
        request = service.spreadsheets().values().batchUpdate(spreadsheetId=spreadsheet_id, body=body)
        response = request.execute()

        print("Values written to the sheet successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")












# Define a variable to store the selected file path
selected_file_path = ""

# Function to browse for a file
def browse_file():
    global selected_file_path
    selected_file_path = filedialog.askopenfilename()
    log_text.insert(tk.END, f"Selected file: {selected_file_path}\n")

# Define a function to perform the entire process

missing_students = []
def process_file():
    global missing_students
    today = datetime.datetime.today()

    # Calculate the date of the next Monday
    days_until_monday = (7 - today.weekday()) % 7
    next_monday = today + datetime.timedelta(days=days_until_monday)
    '''This is a string of the date on monday'''
    monday_date = next_monday.strftime('%Y-%m-%d')


    

    global selected_file_path
    if not selected_file_path:
        log_text.insert(tk.END, "Please select a file first.\n")
        return
    


    log_text.insert(tk.END, "Getting TXT file location\n")

    dictionary = read_config_file('C:/Users/preet/Documents/MATIC SOLUTIONS/MATHNASIUM/system_files/text_file.txt')
    appoint_path = dictionary['appointments folder']
    student_delete_path = dictionary['student list folder']
    template_path = dictionary['template_path']
    

    log_text.insert(tk.END, "Done!\n")



    
    log_text.insert(tk.END, "Cleaning CSV\n")

    cleaning_path = student_delete_path+monday_date+'.xlsx'
    create_appointments_path = appoint_path+monday_date+'.xlsx'

    convert(selected_file_path,cleaning_path)
    date_time_convert(cleaning_path)

    delete_bad_columns(cleaning_path)
    fix_intake(cleaning_path)
    no_student_name(cleaning_path)
    home_online(cleaning_path)

    log_text.insert(tk.END, "Done!\n")



    
    log_text.insert(tk.END, "Creating Excel\n")

    copy_template(create_appointments_path,template_path)
    create_schedule(cleaning_path,create_appointments_path)
    delete_file_excel(cleaning_path)

    log_text.insert(tk.END, "Done!\n")



    
    


    for student_info in missing_students:
        result_text.insert(tk.END, student_info + '\n')


def upload_to_google_drive():
    # Calculate the date of the next Monday
    today = datetime.datetime.today()
    days_until_monday = (7 - today.weekday()) % 7
    next_monday = today + datetime.timedelta(days=days_until_monday)
    '''This is a string of the date on monday'''
    monday_date = next_monday.strftime('%Y-%m-%d')


    dictionary = read_config_file('path')
    service_account_key_file = dictionary['service_account_key_file']
    folder_id = dictionary['folder_id']
    template_id_drive = dictionary['template_id']
    appoint_path = dictionary['appointments folder']
    create_appointments_path = appoint_path+monday_date+'.xlsx'


    log_text.insert(tk.END, "Uploading to Google Drive\n")
    excel_upload_file_id = copy_excel_to_sheets(service_account_key_file,folder_id,create_appointments_path,monday_date)
    sheet_file_id = convert_excel_to_sheets(service_account_key_file,excel_upload_file_id,monday_date)
    delete_file(service_account_key_file, excel_upload_file_id)
    log_text.insert(tk.END, "Done!\n")

    log_text.insert(tk.END, "Copying Template\n")
    copy_column_a(service_account_key_file,template_id_drive,sheet_file_id)
    write_values_to_sheet(service_account_key_file, sheet_file_id,next_monday)
    log_text.insert(tk.END, "Done!\n")
    



# Create the main window
root = tk.Tk()
root.title("SCHEDULE MAKER")
root.geometry("700x450")

label = tk.Label(root, text="Select a file:")
label.pack()
# Create a "Browse" button
browse_button = tk.Button(root, text="Browse", command=browse_file)
browse_button.pack()

# Create a Text widget for displaying log messages
log_text = tk.Text(root, height=10, width=60)
log_text.pack()

# Create another Text widget for displaying the result or additional information
result_text = tk.Text(root, height=10, width=60)
result_text.pack()

# Create a button to trigger the entire process
process_button = tk.Button(root, text="Process File", command=process_file)
process_button.pack()

upload_button = tk.Button(root, text="Upload to Google Drive", command=upload_to_google_drive)
upload_button.pack()

# Start the GUI event loop
root.mainloop()

