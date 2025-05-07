from openpyxl import Workbook
from openpyxl import load_workbook

def fix_intake():

    wb = load_workbook(filename = 'appointmentsReport.xlsx')
    ws = wb['appointmentsReport']

    maxr = ws.max_row
    maxc = ws.max_column

    for r in range(1,maxr+1):

        y = ws['E'+str(r)].value
        y = y.replace('Intake form:\n','')
        y = y.replace('Student Name:','')
        y = y.replace('Student #2: ,\n','')
        y = y.replace('Student #2:','')
        y = y.replace('Student #3:','')
        y = y.replace('Student #2 Name (if applicable): ,','')
        y = y.replace('Student #3 Name (if applicable): ','')
        y = y.replace(',\n','')

        ws['E'+str(r)].value = y

    wb.save('appointmentsReport.xlsx')



fix_intake()
