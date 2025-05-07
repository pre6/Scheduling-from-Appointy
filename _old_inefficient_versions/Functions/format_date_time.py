from openpyxl import Workbook
from openpyxl import load_workbook
from datetime import datetime, time, timedelta

def format_date_time():
    wb = load_workbook(filename = 'appointmentsReport.xlsx')
    ws = wb['appointmentsReport']

    ws.insert_cols(3)

    maxr = ws.max_row

    for i in range(1,maxr+1):
        column = 'A'
        row  = str(i)
        date = ws[column + row].value.date()
        time = ws['B'+row].value
        new = datetime.combine(date,time)

        ws['C' + row].value = new

    ws.delete_cols(1, 2)
    wb.save('appointmentsReport.xlsx')

format_date_time()
