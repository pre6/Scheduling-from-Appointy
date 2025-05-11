
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

# Load both workbooks
wb1 = load_workbook("2025-05-12.xlsx")
wb2 = load_workbook("Template_SCHOOL_YEAR.xlsx")

# Select the first sheet from each workbook (you can change this)
ws1 = wb1.active
ws2 = wb2.active

# Copy first two columns (A and B)
for row in range(1, ws2.max_row + 1):
    for col in [2]:  # Columns A and B
        source_cell = ws2.cell(row=row, column=col)
        target_cell = ws1.cell(row=row, column=col)

        # Copy value
        target_cell.value = source_cell.value

        # Copy style
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = copy(source_cell.number_format)
            target_cell.protection = copy(source_cell.protection)

# Save changes directly to workbook1.xlsx
wb1.save("2025-05-12.xlsx")
