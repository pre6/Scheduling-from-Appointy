
# Python program to explain os.remove() method

# importing os module
import os

# File name
file = 'appointmentsReport.csv'

# File location
location = "C:/Users/Centre Director/Downloads/"

# Path
path = os.path.join(location, file)

# Remove the file
# 'file.txt'
os.remove(path)




















'''

import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook


df = pd.read_csv("C:/Users/Centre Director/Downloads/appointmentsReport.csv")

wb = Workbook()
f = "C:/Users/Centre Director/Desktop/appointmentsReport.xlsx"
wb.save(f)


df.to_excel(f, index=None, header=True)



wb_1 = load_workbook(filename = 'appointmentsReport.xlsx')
ws = wb_1['Sheet1']

print(ws['A3'].value)
print(type(ws['A3'].value))
'''
