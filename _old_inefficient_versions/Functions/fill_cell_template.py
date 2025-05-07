from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from find_cell_template import get_cell, split_list
from create_list import create_list
from find_cell_template import get_cells_ol
import datetime


x = create_list()

y = split_list(x[1])

cells = []
i = 0

for item in y:
    f = get_cells_ol(item,i)
    cells.append(f)
    i += 1


def fill_cell_ol(y):

    wb = load_workbook(filename = 'Template.xlsx')
    ws = wb['Week']

    fill = PatternFill(start_color='EA9999', fill_type='solid')

    for day in y:
        for dict in day:
            col = list(dict.values())[0][1]
            row = list(dict.values())[0][2]

            for item in dict:
                student = dict[item][0]
                ws[col + str(row)].value =  student
                ws[col + str(row)].fill = fill
                ws[col + str(row+1)].fill = fill
                ws[col + str(row+2)].fill = fill
                ws[col + str(row+3)].fill = fill
                row = row + 4

    wb.save('Template.xlsx')

print(cells)
fill_cell_ol(cells)


def fill_cell(y):
    wb = load_workbook(filename = 'Template.xlsx')
    ws = wb['Week']

    first = PatternFill(start_color='FCE5CD', fill_type='solid')
    second = PatternFill(start_color='D9EAD3', fill_type='solid')
    third = PatternFill(start_color='9FC5E8', fill_type='solid')
    fourth = PatternFill(start_color='EAD1DC', fill_type='solid')
    fifth = PatternFill(start_color='B4A7D6', fill_type='solid')
    sixth = PatternFill(start_color='FFD966', fill_type='solid')
    seventh = PatternFill(start_color='FFD966', fill_type='solid')
    eighth = PatternFill(start_color='FFD966', fill_type='solid')
    ninth = PatternFill(start_color='FFD966', fill_type='solid')

    colours = [first,second, third, fourth, fifth,sixth,seventh,eighth,ninth]

    others = PatternFill(start_color='FFD966', fill_type='solid')

    other_times = []
    other_times.append(datetime.time(hour = 10))
    other_times.append(datetime.time(hour = 10,minute = 15))
    other_times.append(datetime.time(hour = 10,minute = 30))
    other_times.append(datetime.time(hour = 15,minute = 30))
    other_times.append(datetime.time(hour = 15, minute = 45))
    other_times.append(datetime.time(hour = 16))



    for day in y:
        i=0
        h= 0

        for dict in day:
            col = list(dict.values())[0][1]
            row = list(dict.values())[0][2]

            if list(dict.keys())[0].time() not in other_times:
                fill = others
            elif h != list(dict.keys())[0].time():
                i = 0
                fill = colours[i]
            else:
                fill = colours[i]

            for item in dict:
                student = dict[item][0]
                ws[col + str(row)].value =  student
                ws[col + str(row)].fill = fill
                ws[col + str(row+1)].fill = fill
                ws[col + str(row+2)].fill = fill
                ws[col + str(row+3)].fill = fill
                row = row + 4
            i+=1
            h = list(dict.keys())[0].time()
    wb.save('Template.xlsx')

#fill_cell(cells)



'''
ws['C' + row].value = new

ws['G7'].fill = redFill
'''
